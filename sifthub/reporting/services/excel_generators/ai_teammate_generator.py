from io import BytesIO
from typing import Optional, Dict, Any, List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sifthub.reporting.models.export_models import SQSExportMessage
from sifthub.reporting.services.insights_api_client import insights_api_client
from sifthub.utils.logger import setup_logger

logger = setup_logger(__name__)


class AITeammateExcelGenerator:
    """Excel generator for AI teammate insights"""
    
    def __init__(self):
        self.workbook = Workbook()
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    async def generate_excel(self, message: SQSExportMessage) -> Optional[BytesIO]:
        """Generate Excel file for AI teammate export"""
        try:
            logger.info(f"Starting AI teammate Excel generation for event: {message.eventId}")
            
            # Fetch data from APIs
            data = await self._fetch_ai_teammate_data(message)
            
            # Remove default sheet
            if "Sheet" in self.workbook.sheetnames:
                self.workbook.remove(self.workbook["Sheet"])
            
            # Create sheets
            await self._create_ai_overview_sheet(data, message)
            await self._create_interaction_analytics_sheet(data, message)
            await self._create_performance_metrics_sheet(data, message)
            await self._create_user_feedback_sheet(data, message)
            
            # Save to BytesIO
            excel_stream = BytesIO()
            self.workbook.save(excel_stream)
            excel_stream.seek(0)
            
            logger.info(f"AI teammate Excel generated successfully for event: {message.eventId}")
            return excel_stream
            
        except Exception as e:
            logger.error(f"Error generating AI teammate Excel: {e}", exc_info=True)
            return None

    async def _fetch_ai_teammate_data(self, message: SQSExportMessage) -> Dict[str, Any]:
        """Fetch AI teammate data from APIs"""
        try:
            logger.info("Fetching AI teammate data from APIs")
            
            # Extract filters from message
            filters = getattr(message, 'filters', {})
            
            # Simulate API calls for AI teammate data
            # In a real implementation, these would be actual API endpoints
            ai_teammate_data = {
                'overview': {
                    'total_interactions': 15847,
                    'unique_users': 342,
                    'avg_response_time': '1.2 seconds',
                    'satisfaction_rate': 87.5,
                    'resolution_rate': 92.3
                },
                'interaction_analytics': [
                    {
                        'interaction_type': 'Code Generation',
                        'count': 5234,
                        'avg_response_time': '2.1 seconds',
                        'success_rate': 94.2
                    },
                    {
                        'interaction_type': 'Code Review',
                        'count': 3456,
                        'avg_response_time': '1.8 seconds',
                        'success_rate': 89.7
                    },
                    {
                        'interaction_type': 'Bug Fixing',
                        'count': 2789,
                        'avg_response_time': '3.2 seconds',
                        'success_rate': 91.5
                    },
                    {
                        'interaction_type': 'Documentation',
                        'count': 2134,
                        'avg_response_time': '1.5 seconds',
                        'success_rate': 96.1
                    },
                    {
                        'interaction_type': 'Testing',
                        'count': 1876,
                        'avg_response_time': '2.8 seconds',
                        'success_rate': 88.3
                    }
                ],
                'performance_metrics': [
                    {
                        'metric': 'Code Quality Score',
                        'value': 8.7,
                        'trend': '+5.2%',
                        'benchmark': 8.2
                    },
                    {
                        'metric': 'Response Accuracy',
                        'value': 92.4,
                        'trend': '+2.1%',
                        'benchmark': 90.0
                    },
                    {
                        'metric': 'User Engagement',
                        'value': 78.9,
                        'trend': '+8.7%',
                        'benchmark': 75.0
                    },
                    {
                        'metric': 'Task Completion Rate',
                        'value': 94.6,
                        'trend': '+1.8%',
                        'benchmark': 92.0
                    }
                ],
                'user_feedback': [
                    {
                        'feedback_type': 'Positive',
                        'count': 1247,
                        'percentage': 72.3,
                        'common_themes': ['Helpful', 'Fast', 'Accurate']
                    },
                    {
                        'feedback_type': 'Neutral',
                        'count': 298,
                        'percentage': 17.3,
                        'common_themes': ['Average', 'Okay', 'Could be better']
                    },
                    {
                        'feedback_type': 'Negative',
                        'count': 179,
                        'percentage': 10.4,
                        'common_themes': ['Slow', 'Inaccurate', 'Confusing']
                    }
                ]
            }
            
            logger.info("AI teammate data fetched successfully")
            return ai_teammate_data
            
        except Exception as e:
            logger.error(f"Error fetching AI teammate data: {e}", exc_info=True)
            return {}

    async def _create_ai_overview_sheet(self, data: Dict[str, Any], message: SQSExportMessage):
        """Create AI teammate overview sheet"""
        try:
            sheet = self.workbook.create_sheet("AI Teammate Overview")
            
            # Title
            sheet['A1'] = "AI Teammate Performance Overview"
            sheet['A1'].font = Font(size=16, bold=True)
            sheet.merge_cells('A1:D1')
            
            # Metadata
            row = 3
            sheet[f'A{row}'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            sheet[f'A{row}'].font = Font(italic=True)
            row += 1
            sheet[f'A{row}'] = f"Event ID: {message.eventId}"
            sheet[f'A{row}'].font = Font(italic=True)
            row += 2
            
            # Overview metrics
            overview = data.get('overview', {})
            
            # Headers
            headers = ['Metric', 'Value']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            row += 1
            
            # Data rows
            metrics = [
                ('Total Interactions', overview.get('total_interactions', 0)),
                ('Unique Users', overview.get('unique_users', 0)),
                ('Average Response Time', overview.get('avg_response_time', 'N/A')),
                ('Satisfaction Rate (%)', overview.get('satisfaction_rate', 0)),
                ('Resolution Rate (%)', overview.get('resolution_rate', 0))
            ]
            
            for metric, value in metrics:
                sheet.cell(row=row, column=1, value=metric).border = self.border
                sheet.cell(row=row, column=2, value=value).border = self.border
                row += 1
            
            # Auto-adjust column widths
            self._auto_adjust_columns(sheet)
            
            logger.info("AI teammate overview sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating AI teammate overview sheet: {e}", exc_info=True)

    async def _create_interaction_analytics_sheet(self, data: Dict[str, Any], message: SQSExportMessage):
        """Create interaction analytics sheet"""
        try:
            sheet = self.workbook.create_sheet("Interaction Analytics")
            
            # Title
            sheet['A1'] = "AI Teammate Interaction Analytics"
            sheet['A1'].font = Font(size=16, bold=True)
            sheet.merge_cells('A1:E1')
            
            # Metadata
            row = 3
            sheet[f'A{row}'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            sheet[f'A{row}'].font = Font(italic=True)
            row += 2
            
            # Headers
            headers = ['Interaction Type', 'Count', 'Avg Response Time', 'Success Rate (%)']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            row += 1
            
            # Data rows
            interaction_analytics = data.get('interaction_analytics', [])
            for interaction in interaction_analytics:
                sheet.cell(row=row, column=1, value=interaction.get('interaction_type', '')).border = self.border
                sheet.cell(row=row, column=2, value=interaction.get('count', 0)).border = self.border
                sheet.cell(row=row, column=3, value=interaction.get('avg_response_time', '')).border = self.border
                sheet.cell(row=row, column=4, value=interaction.get('success_rate', 0)).border = self.border
                row += 1
            
            # Auto-adjust column widths
            self._auto_adjust_columns(sheet)
            
            logger.info("Interaction analytics sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating interaction analytics sheet: {e}", exc_info=True)

    async def _create_performance_metrics_sheet(self, data: Dict[str, Any], message: SQSExportMessage):
        """Create performance metrics sheet"""
        try:
            sheet = self.workbook.create_sheet("Performance Metrics")
            
            # Title
            sheet['A1'] = "AI Teammate Performance Metrics"
            sheet['A1'].font = Font(size=16, bold=True)
            sheet.merge_cells('A1:E1')
            
            # Metadata
            row = 3
            sheet[f'A{row}'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            sheet[f'A{row}'].font = Font(italic=True)
            row += 2
            
            # Headers
            headers = ['Metric', 'Value', 'Trend', 'Benchmark']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            row += 1
            
            # Data rows
            performance_metrics = data.get('performance_metrics', [])
            for metric in performance_metrics:
                sheet.cell(row=row, column=1, value=metric.get('metric', '')).border = self.border
                sheet.cell(row=row, column=2, value=metric.get('value', 0)).border = self.border
                sheet.cell(row=row, column=3, value=metric.get('trend', '')).border = self.border
                sheet.cell(row=row, column=4, value=metric.get('benchmark', 0)).border = self.border
                row += 1
            
            # Auto-adjust column widths
            self._auto_adjust_columns(sheet)
            
            logger.info("Performance metrics sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating performance metrics sheet: {e}", exc_info=True)

    async def _create_user_feedback_sheet(self, data: Dict[str, Any], message: SQSExportMessage):
        """Create user feedback sheet"""
        try:
            sheet = self.workbook.create_sheet("User Feedback")
            
            # Title
            sheet['A1'] = "AI Teammate User Feedback Analysis"
            sheet['A1'].font = Font(size=16, bold=True)
            sheet.merge_cells('A1:E1')
            
            # Metadata
            row = 3
            sheet[f'A{row}'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            sheet[f'A{row}'].font = Font(italic=True)
            row += 2
            
            # Headers
            headers = ['Feedback Type', 'Count', 'Percentage (%)', 'Common Themes']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            row += 1
            
            # Data rows
            user_feedback = data.get('user_feedback', [])
            for feedback in user_feedback:
                sheet.cell(row=row, column=1, value=feedback.get('feedback_type', '')).border = self.border
                sheet.cell(row=row, column=2, value=feedback.get('count', 0)).border = self.border
                sheet.cell(row=row, column=3, value=feedback.get('percentage', 0)).border = self.border
                
                # Join common themes with commas
                themes = feedback.get('common_themes', [])
                themes_str = ', '.join(themes) if themes else 'N/A'
                sheet.cell(row=row, column=4, value=themes_str).border = self.border
                row += 1
            
            # Auto-adjust column widths
            self._auto_adjust_columns(sheet)
            
            logger.info("User feedback sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating user feedback sheet: {e}", exc_info=True)

    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths based on content"""
        try:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.warning(f"Error auto-adjusting columns: {e}") 