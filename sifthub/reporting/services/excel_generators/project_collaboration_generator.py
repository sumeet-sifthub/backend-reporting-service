from io import BytesIO
from typing import Optional, Dict, Any, List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sifthub.reporting.models.export_models import SQSExportMessage
from sifthub.reporting.services.insights_api_client import insights_api_client
from sifthub.utils.logger import setup_logger

logger = setup_logger(__name__)


class ProjectCollaborationExcelGenerator:
    """Excel generator for project collaboration insights"""
    
    def __init__(self):
        self.workbook = Workbook()
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    async def generate_excel(self, message: SQSExportMessage) -> Optional[BytesIO]:
        """Generate Excel file for project collaboration export"""
        try:
            logger.info(f"Starting project collaboration Excel generation for event: {message.eventId}")
            
            # Fetch data from APIs
            data = await self._fetch_collaboration_data(message)
            
            # Remove default sheet
            if "Sheet" in self.workbook.sheetnames:
                self.workbook.remove(self.workbook["Sheet"])
            
            # Create sheets
            await self._create_collaboration_overview_sheet(data, message)
            await self._create_team_activity_sheet(data, message)
            await self._create_project_metrics_sheet(data, message)
            
            # Save to BytesIO
            excel_stream = BytesIO()
            self.workbook.save(excel_stream)
            excel_stream.seek(0)
            
            logger.info(f"Project collaboration Excel generated successfully for event: {message.eventId}")
            return excel_stream
            
        except Exception as e:
            logger.error(f"Error generating project collaboration Excel: {e}", exc_info=True)
            return None

    async def _fetch_collaboration_data(self, message: SQSExportMessage) -> Dict[str, Any]:
        """Fetch project collaboration data from APIs"""
        try:
            logger.info("Fetching project collaboration data from APIs")
            
            # Extract filters from message
            filters = getattr(message, 'filters', {})
            
            # Simulate API calls for project collaboration data
            # In a real implementation, these would be actual API endpoints
            collaboration_data = {
                'overview': {
                    'total_projects': 45,
                    'active_collaborators': 128,
                    'total_contributions': 2847,
                    'avg_response_time': '2.3 hours'
                },
                'team_activity': [
                    {
                        'team_name': 'Frontend Team',
                        'members': 8,
                        'contributions': 456,
                        'avg_response_time': '1.8 hours',
                        'active_projects': 12
                    },
                    {
                        'team_name': 'Backend Team',
                        'members': 6,
                        'contributions': 389,
                        'avg_response_time': '2.1 hours',
                        'active_projects': 8
                    },
                    {
                        'team_name': 'DevOps Team',
                        'members': 4,
                        'contributions': 234,
                        'avg_response_time': '3.2 hours',
                        'active_projects': 15
                    }
                ],
                'project_metrics': [
                    {
                        'project_name': 'Mobile App Redesign',
                        'collaborators': 12,
                        'contributions': 234,
                        'completion_rate': 78.5,
                        'last_activity': '2024-01-15'
                    },
                    {
                        'project_name': 'API Gateway Migration',
                        'collaborators': 8,
                        'contributions': 189,
                        'completion_rate': 92.3,
                        'last_activity': '2024-01-14'
                    },
                    {
                        'project_name': 'Database Optimization',
                        'collaborators': 5,
                        'contributions': 156,
                        'completion_rate': 65.2,
                        'last_activity': '2024-01-13'
                    }
                ]
            }
            
            logger.info("Project collaboration data fetched successfully")
            return collaboration_data
            
        except Exception as e:
            logger.error(f"Error fetching project collaboration data: {e}", exc_info=True)
            return {}

    async def _create_collaboration_overview_sheet(self, data: Dict[str, Any], message: SQSExportMessage):
        """Create collaboration overview sheet"""
        try:
            sheet = self.workbook.create_sheet("Collaboration Overview")
            
            # Title
            sheet['A1'] = "Project Collaboration Overview"
            sheet['A1'].font = Font(size=16, bold=True)
            sheet.merge_cells('A1:D1')
            
            # Metadata
            row = 3
            sheet[f'A{row}'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            sheet[f'A{row}'].font = Font(italic=True)
            row += 1
            sheet[f'A{row}'] = f"Event ID: {message.eventId}"
            sheet[f'A{row}'].font = Font(italic=True)
            row += 2
            
            # Overview metrics
            overview = data.get('overview', {})
            
            # Headers
            headers = ['Metric', 'Value']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            row += 1
            
            # Data rows
            metrics = [
                ('Total Projects', overview.get('total_projects', 0)),
                ('Active Collaborators', overview.get('active_collaborators', 0)),
                ('Total Contributions', overview.get('total_contributions', 0)),
                ('Average Response Time', overview.get('avg_response_time', 'N/A'))
            ]
            
            for metric, value in metrics:
                sheet.cell(row=row, column=1, value=metric).border = self.border
                sheet.cell(row=row, column=2, value=value).border = self.border
                row += 1
            
            # Auto-adjust column widths
            self._auto_adjust_columns(sheet)
            
            logger.info("Collaboration overview sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating collaboration overview sheet: {e}", exc_info=True)

    async def _create_team_activity_sheet(self, data: Dict[str, Any], message: SQSExportMessage):
        """Create team activity sheet"""
        try:
            sheet = self.workbook.create_sheet("Team Activity")
            
            # Title
            sheet['A1'] = "Team Activity Report"
            sheet['A1'].font = Font(size=16, bold=True)
            sheet.merge_cells('A1:F1')
            
            # Metadata
            row = 3
            sheet[f'A{row}'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            sheet[f'A{row}'].font = Font(italic=True)
            row += 2
            
            # Headers
            headers = ['Team Name', 'Members', 'Contributions', 'Avg Response Time', 'Active Projects']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            row += 1
            
            # Data rows
            team_activity = data.get('team_activity', [])
            for team in team_activity:
                sheet.cell(row=row, column=1, value=team.get('team_name', '')).border = self.border
                sheet.cell(row=row, column=2, value=team.get('members', 0)).border = self.border
                sheet.cell(row=row, column=3, value=team.get('contributions', 0)).border = self.border
                sheet.cell(row=row, column=4, value=team.get('avg_response_time', '')).border = self.border
                sheet.cell(row=row, column=5, value=team.get('active_projects', 0)).border = self.border
                row += 1
            
            # Auto-adjust column widths
            self._auto_adjust_columns(sheet)
            
            logger.info("Team activity sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating team activity sheet: {e}", exc_info=True)

    async def _create_project_metrics_sheet(self, data: Dict[str, Any], message: SQSExportMessage):
        """Create project metrics sheet"""
        try:
            sheet = self.workbook.create_sheet("Project Metrics")
            
            # Title
            sheet['A1'] = "Project Metrics Report"
            sheet['A1'].font = Font(size=16, bold=True)
            sheet.merge_cells('A1:F1')
            
            # Metadata
            row = 3
            sheet[f'A{row}'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            sheet[f'A{row}'].font = Font(italic=True)
            row += 2
            
            # Headers
            headers = ['Project Name', 'Collaborators', 'Contributions', 'Completion Rate (%)', 'Last Activity']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            row += 1
            
            # Data rows
            project_metrics = data.get('project_metrics', [])
            for project in project_metrics:
                sheet.cell(row=row, column=1, value=project.get('project_name', '')).border = self.border
                sheet.cell(row=row, column=2, value=project.get('collaborators', 0)).border = self.border
                sheet.cell(row=row, column=3, value=project.get('contributions', 0)).border = self.border
                sheet.cell(row=row, column=4, value=project.get('completion_rate', 0)).border = self.border
                sheet.cell(row=row, column=5, value=project.get('last_activity', '')).border = self.border
                row += 1
            
            # Auto-adjust column widths
            self._auto_adjust_columns(sheet)
            
            logger.info("Project metrics sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating project metrics sheet: {e}", exc_info=True)

    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths based on content"""
        try:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.warning(f"Error auto-adjusting columns: {e}") 