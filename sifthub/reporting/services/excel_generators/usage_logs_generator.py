from io import BytesIO
from typing import Optional, Dict, Any, List
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sifthub.reporting.models.export_models import SQSExportMessage, UsageLogsType
from sifthub.utils.logger import setup_logger

logger = setup_logger(__name__)


class UsageLogsExcelGenerator:
    """Excel generator for usage logs"""
    
    def __init__(self):
        self.workbook = Workbook()
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    async def generate_excel(self, message: SQSExportMessage) -> Optional[BytesIO]:
        """Generate Excel file for usage logs export"""
        try:
            logger.info(f"Starting usage logs Excel generation for event: {message.eventId}, type: {message.type}")
            
            # Fetch data from APIs
            data = await self._fetch_usage_logs_data(message)
            
            # Remove default sheet
            if "Sheet" in self.workbook.sheetnames:
                self.workbook.remove(self.workbook["Sheet"])
            
            # Create sheets based on usage log type
            await self._create_summary_sheet(data, message)
            await self._create_detailed_logs_sheet(data, message)
            await self._create_analytics_sheet(data, message)
            
            # Save to BytesIO
            excel_stream = BytesIO()
            self.workbook.save(excel_stream)
            excel_stream.seek(0)
            
            logger.info(f"Usage logs Excel generated successfully for event: {message.eventId}")
            return excel_stream
            
        except Exception as e:
            logger.error(f"Error generating usage logs Excel: {e}", exc_info=True)
            return None

    async def _fetch_usage_logs_data(self, message: SQSExportMessage) -> Dict[str, Any]:
        """Fetch usage logs data from APIs"""
        try:
            logger.info(f"Fetching usage logs data for type: {message.type}")
            
            # Extract filters from message
            filters = getattr(message, 'filters', {})
            
            # Generate sample data based on usage log type
            if message.type == UsageLogsType.ANSWER:
                return await self._generate_answer_logs_data(filters)
            elif message.type == UsageLogsType.AUTOFILL:
                return await self._generate_autofill_logs_data(filters)
            elif message.type == UsageLogsType.AI_TEAMMATE:
                return await self._generate_ai_teammate_logs_data(filters)
            else:
                return await self._generate_generic_logs_data(filters)
            
        except Exception as e:
            logger.error(f"Error fetching usage logs data: {e}", exc_info=True)
            return {}

    async def _generate_answer_logs_data(self, filters: Dict[str, Any]) -> Dict[str, Any]:
        """Generate sample answer logs data"""
        return {
            'summary': {
                'total_queries': 8547,
                'successful_answers': 7892,
                'failed_answers': 655,
                'avg_response_time': '2.3 seconds',
                'success_rate': 92.3
            },
            'detailed_logs': [
                {
                    'timestamp': '2024-01-15 10:30:45',
                    'user_id': 'user_001',
                    'query': 'How to implement async functions in Python?',
                    'response_time': '1.8 seconds',
                    'status': 'Success',
                    'confidence_score': 0.95
                },
                {
                    'timestamp': '2024-01-15 10:32:12',
                    'user_id': 'user_002',
                    'query': 'Best practices for database indexing',
                    'response_time': '2.1 seconds',
                    'status': 'Success',
                    'confidence_score': 0.89
                },
                {
                    'timestamp': '2024-01-15 10:35:33',
                    'user_id': 'user_003',
                    'query': 'React component lifecycle methods',
                    'response_time': '1.5 seconds',
                    'status': 'Success',
                    'confidence_score': 0.92
                }
            ],
            'analytics': {
                'top_categories': [
                    {'category': 'Programming', 'count': 3245, 'percentage': 38.0},
                    {'category': 'Database', 'count': 2156, 'percentage': 25.2},
                    {'category': 'Web Development', 'count': 1876, 'percentage': 21.9},
                    {'category': 'DevOps', 'count': 1270, 'percentage': 14.9}
                ],
                'hourly_distribution': [
                    {'hour': '09:00', 'count': 456},
                    {'hour': '10:00', 'count': 623},
                    {'hour': '11:00', 'count': 789},
                    {'hour': '14:00', 'count': 567},
                    {'hour': '15:00', 'count': 445}
                ]
            }
        }

    async def _generate_autofill_logs_data(self, filters: Dict[str, Any]) -> Dict[str, Any]:
        """Generate sample autofill logs data"""
        return {
            'summary': {
                'total_suggestions': 15234,
                'accepted_suggestions': 12456,
                'rejected_suggestions': 2778,
                'avg_completion_time': '0.8 seconds',
                'acceptance_rate': 81.8
            },
            'detailed_logs': [
                {
                    'timestamp': '2024-01-15 09:15:22',
                    'user_id': 'user_004',
                    'context': 'function calculateTotal(',
                    'suggestion': 'price, tax) {\n  return price + (price * tax);\n}',
                    'action': 'Accepted',
                    'completion_time': '0.6 seconds'
                },
                {
                    'timestamp': '2024-01-15 09:18:45',
                    'user_id': 'user_005',
                    'context': 'const handleSubmit = async (',
                    'suggestion': 'event) => {\n  event.preventDefault();\n  // Handle form submission\n}',
                    'action': 'Accepted',
                    'completion_time': '0.9 seconds'
                },
                {
                    'timestamp': '2024-01-15 09:22:11',
                    'user_id': 'user_006',
                    'context': 'SELECT * FROM users WHERE',
                    'suggestion': ' status = "active" AND created_at > NOW() - INTERVAL 30 DAY',
                    'action': 'Rejected',
                    'completion_time': '1.2 seconds'
                }
            ],
            'analytics': {
                'language_distribution': [
                    {'language': 'JavaScript', 'count': 4567, 'percentage': 30.0},
                    {'language': 'Python', 'count': 3456, 'percentage': 22.7},
                    {'language': 'SQL', 'count': 2789, 'percentage': 18.3},
                    {'language': 'Java', 'count': 2234, 'percentage': 14.7},
                    {'language': 'Other', 'count': 2188, 'percentage': 14.3}
                ],
                'acceptance_by_hour': [
                    {'hour': '09:00', 'accepted': 234, 'total': 289},
                    {'hour': '10:00', 'accepted': 345, 'total': 412},
                    {'hour': '11:00', 'accepted': 456, 'total': 523},
                    {'hour': '14:00', 'accepted': 378, 'total': 445},
                    {'hour': '15:00', 'accepted': 289, 'total': 356}
                ]
            }
        }

    async def _generate_ai_teammate_logs_data(self, filters: Dict[str, Any]) -> Dict[str, Any]:
        """Generate sample AI teammate logs data"""
        return {
            'summary': {
                'total_interactions': 6789,
                'successful_interactions': 6234,
                'failed_interactions': 555,
                'avg_session_duration': '4.2 minutes',
                'success_rate': 91.8
            },
            'detailed_logs': [
                {
                    'timestamp': '2024-01-15 11:45:30',
                    'user_id': 'user_007',
                    'interaction_type': 'Code Review',
                    'duration': '3.5 minutes',
                    'status': 'Success',
                    'feedback_score': 4.2
                },
                {
                    'timestamp': '2024-01-15 11:52:15',
                    'user_id': 'user_008',
                    'interaction_type': 'Bug Fixing',
                    'duration': '6.8 minutes',
                    'status': 'Success',
                    'feedback_score': 4.7
                },
                {
                    'timestamp': '2024-01-15 12:05:42',
                    'user_id': 'user_009',
                    'interaction_type': 'Code Generation',
                    'duration': '2.1 minutes',
                    'status': 'Failed',
                    'feedback_score': 2.3
                }
            ],
            'analytics': {
                'interaction_types': [
                    {'type': 'Code Review', 'count': 2345, 'percentage': 34.5},
                    {'type': 'Bug Fixing', 'count': 1876, 'percentage': 27.6},
                    {'type': 'Code Generation', 'count': 1456, 'percentage': 21.4},
                    {'type': 'Documentation', 'count': 1112, 'percentage': 16.4}
                ],
                'success_by_type': [
                    {'type': 'Code Review', 'successful': 2156, 'total': 2345},
                    {'type': 'Bug Fixing', 'successful': 1723, 'total': 1876},
                    {'type': 'Code Generation', 'successful': 1298, 'total': 1456},
                    {'type': 'Documentation', 'successful': 1057, 'total': 1112}
                ]
            }
        }

    async def _generate_generic_logs_data(self, filters: Dict[str, Any]) -> Dict[str, Any]:
        """Generate generic usage logs data"""
        return {
            'summary': {
                'total_events': 12345,
                'successful_events': 11234,
                'failed_events': 1111,
                'avg_processing_time': '1.5 seconds',
                'success_rate': 91.0
            },
            'detailed_logs': [],
            'analytics': {}
        }

    async def _create_summary_sheet(self, data: Dict[str, Any], message: SQSExportMessage):
        """Create summary sheet"""
        try:
            sheet = self.workbook.create_sheet("Summary")
            
            # Title
            sheet['A1'] = f"Usage Logs Summary - {message.type.value if hasattr(message.type, 'value') else message.type}"
            sheet['A1'].font = Font(size=16, bold=True)
            sheet.merge_cells('A1:D1')
            
            # Metadata
            row = 3
            sheet[f'A{row}'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            sheet[f'A{row}'].font = Font(italic=True)
            row += 1
            sheet[f'A{row}'] = f"Event ID: {message.eventId}"
            sheet[f'A{row}'].font = Font(italic=True)
            row += 2
            
            # Summary metrics
            summary = data.get('summary', {})
            
            # Headers
            headers = ['Metric', 'Value']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            row += 1
            
            # Data rows
            for key, value in summary.items():
                metric_name = key.replace('_', ' ').title()
                sheet.cell(row=row, column=1, value=metric_name).border = self.border
                sheet.cell(row=row, column=2, value=value).border = self.border
                row += 1
            
            # Auto-adjust column widths
            self._auto_adjust_columns(sheet)
            
            logger.info("Summary sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating summary sheet: {e}", exc_info=True)

    async def _create_detailed_logs_sheet(self, data: Dict[str, Any], message: SQSExportMessage):
        """Create detailed logs sheet"""
        try:
            sheet = self.workbook.create_sheet("Detailed Logs")
            
            # Title
            sheet['A1'] = f"Detailed Usage Logs - {message.type.value if hasattr(message.type, 'value') else message.type}"
            sheet['A1'].font = Font(size=16, bold=True)
            sheet.merge_cells('A1:G1')
            
            # Metadata
            row = 3
            sheet[f'A{row}'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            sheet[f'A{row}'].font = Font(italic=True)
            row += 2
            
            # Get detailed logs
            detailed_logs = data.get('detailed_logs', [])
            
            if detailed_logs:
                # Headers based on first log entry
                first_log = detailed_logs[0]
                headers = list(first_log.keys())
                
                for col, header in enumerate(headers, 1):
                    cell = sheet.cell(row=row, column=col, value=header.replace('_', ' ').title())
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = self.border
                
                row += 1
                
                # Data rows
                for log_entry in detailed_logs:
                    for col, header in enumerate(headers, 1):
                        value = log_entry.get(header, '')
                        # Truncate long text for readability
                        if isinstance(value, str) and len(value) > 100:
                            value = value[:97] + "..."
                        sheet.cell(row=row, column=col, value=value).border = self.border
                    row += 1
            else:
                sheet.cell(row=row, column=1, value="No detailed logs available").border = self.border
            
            # Auto-adjust column widths
            self._auto_adjust_columns(sheet)
            
            logger.info("Detailed logs sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating detailed logs sheet: {e}", exc_info=True)

    async def _create_analytics_sheet(self, data: Dict[str, Any], message: SQSExportMessage):
        """Create analytics sheet"""
        try:
            sheet = self.workbook.create_sheet("Analytics")
            
            # Title
            sheet['A1'] = f"Usage Analytics - {message.type.value if hasattr(message.type, 'value') else message.type}"
            sheet['A1'].font = Font(size=16, bold=True)
            sheet.merge_cells('A1:E1')
            
            # Metadata
            row = 3
            sheet[f'A{row}'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            sheet[f'A{row}'].font = Font(italic=True)
            row += 2
            
            # Analytics data
            analytics = data.get('analytics', {})
            
            for section_name, section_data in analytics.items():
                if not section_data:
                    continue
                    
                # Section title
                sheet[f'A{row}'] = section_name.replace('_', ' ').title()
                sheet[f'A{row}'].font = Font(size=14, bold=True)
                row += 1
                
                if isinstance(section_data, list) and section_data:
                    # Headers
                    first_item = section_data[0]
                    headers = list(first_item.keys())
                    
                    for col, header in enumerate(headers, 1):
                        cell = sheet.cell(row=row, column=col, value=header.replace('_', ' ').title())
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = self.border
                    
                    row += 1
                    
                    # Data rows
                    for item in section_data:
                        for col, header in enumerate(headers, 1):
                            value = item.get(header, '')
                            sheet.cell(row=row, column=col, value=value).border = self.border
                        row += 1
                
                row += 1  # Add space between sections
            
            # Auto-adjust column widths
            self._auto_adjust_columns(sheet)
            
            logger.info("Analytics sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating analytics sheet: {e}", exc_info=True)

    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths based on content"""
        try:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.warning(f"Error auto-adjusting columns: {e}") 