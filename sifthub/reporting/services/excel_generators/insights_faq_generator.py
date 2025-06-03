from io import BytesIO
from typing import Optional, Dict, Any, List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from sifthub.reporting.models.export_models import SQSExportMessage, FilterConditions
from sifthub.reporting.services.insights_api_client import insights_api_client
from sifthub.utils.logger import setup_logger

logger = setup_logger(__name__)


class InsightsFAQExcelGenerator:
    """Generator for Insights FAQ Excel reports"""

    def __init__(self):
        self.workbook = None
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")

    async def generate_excel(self, message: SQSExportMessage) -> Optional[BytesIO]:
        """Generate Excel file for insights FAQ export"""
        try:
            logger.info(f"Generating insights FAQ Excel for event: {message.eventId}")
            
            # Fetch data from APIs
            data = await self._fetch_data(message)
            if not data:
                logger.error("Failed to fetch data for Excel generation")
                return None

            # Create workbook
            self.workbook = Workbook()
            
            # Generate date range for sheet names
            date_range = self._extract_date_range(message.pageFilter)
            sheet_suffix = self._get_sheet_suffix(message.filter)
            
            # Create sheets
            await self._create_category_sheet(data, date_range, sheet_suffix)
            await self._create_subcategory_sheet(data, date_range, sheet_suffix)
            await self._create_top_questions_sheet(data, date_range, sheet_suffix)
            
            # Remove default sheet
            if "Sheet" in self.workbook.sheetnames:
                self.workbook.remove(self.workbook["Sheet"])
            
            # Save to BytesIO
            stream = BytesIO()
            self.workbook.save(stream)
            stream.seek(0)
            
            logger.info(f"Successfully generated Excel file for event: {message.eventId}")
            return stream
            
        except Exception as e:
            logger.error(f"Error generating Excel file: {e}", exc_info=True)
            return None

    async def _fetch_data(self, message: SQSExportMessage) -> Optional[Dict[str, Any]]:
        """Fetch all required data from APIs"""
        try:
            # Fetch info cards data
            info_cards = await insights_api_client.get_info_cards(message.pageFilter)
            if not info_cards:
                logger.error("Failed to fetch info cards data")
                return None

            # Fetch category distribution
            category_distribution = await insights_api_client.get_category_distribution(
                message.filter, message.pageFilter
            )
            if not category_distribution:
                logger.error("Failed to fetch category distribution")
                return None

            # Fetch subcategory data for each category
            subcategory_data = {}
            for category in category_distribution.category:
                subcategory_response = await insights_api_client.get_subcategory_distribution(
                    category.id, message.filter, message.pageFilter
                )
                if subcategory_response:
                    subcategory_data[category.id] = {
                        "category_name": category.category,
                        "subcategories": subcategory_response.subCategory
                    }

            # Fetch top questions
            top_questions = await insights_api_client.get_all_top_questions(
                message.filter, message.pageFilter
            )

            return {
                "info_cards": info_cards,
                "category_distribution": category_distribution,
                "subcategory_data": subcategory_data,
                "top_questions": top_questions
            }

        except Exception as e:
            logger.error(f"Error fetching data: {e}", exc_info=True)
            return None

    async def _create_category_sheet(self, data: Dict[str, Any], date_range: str, sheet_suffix: str):
        """Create the top question categories sheet"""
        try:
            sheet_name = f"Top question categories - {sheet_suffix}"
            ws = self.workbook.create_sheet(title=sheet_name)
            
            # Add title and date range
            ws.merge_cells('A1:E1')
            ws['A1'] = f"Frequently Asked Questions Report - {date_range}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Add headers
            headers = ["Category", "Frequency (Questions asked)", "Distribution", "Trend"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment

            # Add data
            info_cards = data["info_cards"]
            category_distribution = data["category_distribution"]
            
            # Calculate base count for frequency calculation
            base_count = self._get_base_count(info_cards, sheet_suffix)
            
            row = 4
            for category in category_distribution.category:
                frequency = int(base_count * (category.distribution / 100))
                trend_symbol = "▲" if category.direction == "INCREASING" else "▼"
                trend_text = f"{trend_symbol} {abs(category.trend):.1f}%"
                
                ws.cell(row=row, column=1, value=category.category)
                ws.cell(row=row, column=2, value=frequency)
                ws.cell(row=row, column=3, value=f"{category.distribution:.1f}%")
                ws.cell(row=row, column=4, value=trend_text)
                row += 1

            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating category sheet: {e}", exc_info=True)

    async def _create_subcategory_sheet(self, data: Dict[str, Any], date_range: str, sheet_suffix: str):
        """Create the detailed category breakdown sheet"""
        try:
            sheet_name = f"Detailed category breakdown - {sheet_suffix}"
            ws = self.workbook.create_sheet(title=sheet_name)
            
            # Add title and date range
            ws.merge_cells('A1:F1')
            ws['A1'] = f"Frequently Asked Questions Report - {date_range}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Add headers
            headers = ["Subcategory", "Parent category", "Frequency (Questions asked)", "Distrubution", "Trend"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment

            # Add data
            info_cards = data["info_cards"]
            subcategory_data = data["subcategory_data"]
            
            # Calculate base count for frequency calculation
            base_count = self._get_base_count(info_cards, sheet_suffix)
            
            row = 4
            for category_id, category_info in subcategory_data.items():
                parent_category = category_info["category_name"]
                for subcategory in category_info["subcategories"]:
                    frequency = int(base_count * (subcategory.distribution / 100))
                    trend_symbol = "▲" if subcategory.direction == "INCREASING" else "▼"
                    trend_text = f"{trend_symbol} {abs(subcategory.trend):.1f}%"
                    
                    ws.cell(row=row, column=1, value=subcategory.subCategory)
                    ws.cell(row=row, column=2, value=parent_category)
                    ws.cell(row=row, column=3, value=frequency)
                    ws.cell(row=row, column=4, value=f"{subcategory.distribution:.1f}%")
                    ws.cell(row=row, column=5, value=trend_text)
                    row += 1

            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating subcategory sheet: {e}", exc_info=True)

    async def _create_top_questions_sheet(self, data: Dict[str, Any], date_range: str, sheet_suffix: str):
        """Create the top asked questions sheet"""
        try:
            sheet_name = f"Top asked questions - {sheet_suffix}"
            ws = self.workbook.create_sheet(title=sheet_name)
            
            # Add title and date range
            ws.merge_cells('A1:C1')
            ws['A1'] = f"Frequently Asked Questions Report - {date_range}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Add headers
            headers = ["Question", "Frequency (Questions asked)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment

            # Add data
            top_questions = data["top_questions"]
            
            row = 4
            for question_data in top_questions:
                ws.cell(row=row, column=1, value=question_data["question"])
                ws.cell(row=row, column=2, value=question_data["frequency"])
                row += 1

            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating top questions sheet: {e}", exc_info=True)

    def _get_base_count(self, info_cards, sheet_suffix: str) -> int:
        """Get base count for frequency calculation based on filter type"""
        try:
            if sheet_suffix == "All":
                return info_cards.totalQuestions["count"]
            elif sheet_suffix == "Answered":
                return info_cards.totalQuestionsAnswered["count"]
            elif sheet_suffix == "Unanswered":
                return info_cards.totalQuestions["count"] - info_cards.totalQuestionsAnswered["count"]
            else:
                return info_cards.totalQuestions["count"]
        except Exception as e:
            logger.error(f"Error getting base count: {e}")
            return 0

    def _get_sheet_suffix(self, filter_conditions: Optional[FilterConditions]) -> str:
        """Determine sheet suffix based on filter conditions"""
        try:
            if not filter_conditions or not filter_conditions.conditions.get("status"):
                return "All"
            
            status_data = filter_conditions.conditions["status"].data
            
            if status_data == "ANSWERED#@#NO_INFORMATION#@#PARTIAL":
                return "All"
            elif status_data == "ANSWERED#@#PARTIAL":
                return "Answered"
            elif status_data == "NO_INFORMATION":
                return "Unanswered"
            else:
                return "All"
                
        except Exception as e:
            logger.error(f"Error determining sheet suffix: {e}")
            return "All"

    def _extract_date_range(self, page_filter: Optional[FilterConditions]) -> str:
        """Extract date range from page filter for display"""
        try:
            if not page_filter or not page_filter.conditions.get("meta.created"):
                return "All Time"
            
            date_data = page_filter.conditions["meta.created"].data
            if "#@#" in date_data:
                start_timestamp, end_timestamp = date_data.split("#@#")
                start_date = datetime.fromtimestamp(int(start_timestamp) / 1000)
                end_date = datetime.fromtimestamp(int(end_timestamp) / 1000)
                return f"{start_date.strftime('%b %d, %Y')} to {end_date.strftime('%b %d, %Y')}"
            
            return "All Time"
            
        except Exception as e:
            logger.error(f"Error extracting date range: {e}")
            return "All Time"

    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths based on content"""
        try:
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.error(f"Error auto-adjusting columns: {e}") 