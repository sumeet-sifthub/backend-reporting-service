import time
from io import BytesIO
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

from sifthub.reporting.models.export_models import SQSExportMessage
from sifthub.reporting.services.insights_analytics_client import InsightsAnalyticsClient
from sifthub.configs.constants import BATCH_SIZE
from sifthub.datastores.document.s3.s3_client import S3Client
from sifthub.utils.logger import setup_logger

logger = setup_logger(__name__)

# Module-level clients
_analytics_client = InsightsAnalyticsClient()
_s3_client = S3Client()


class InsightsFAQExcelGenerator:
    # Excel generator implementing true streaming batch processing
    
    async def generate_excel_streaming(self, message: SQSExportMessage) -> Optional[Dict[str, str]]:
        # Implement the exact flow: Fetch Batch → Write to Excel → Stream to S3 → Repeat
        try:
            logger.info(f"Starting streaming batch processing for event: {message.eventId}")
            
            # Step 1: Initialize Excel structure and upload to S3
            s3_key = await self._initialize_excel_structure(message)
            if not s3_key:
                logger.error("Failed to initialize Excel structure")
                return None
            
            # Step 2: Process all batches sequentially
            success = await self._process_all_batches_streaming(message, s3_key)
            if not success:
                logger.error("Failed to process all batches")
                return None
            
            # Step 3: Generate presigned URL (after all batching is complete)
            download_url = await _s3_client.generate_presigned_url(s3_key)
            if not download_url:
                logger.error("Failed to generate presigned URL")
                return None
            
            logger.info(f"Successfully completed streaming batch processing for event: {message.eventId}")
            
            # Step 4: Return result for Firebase notification
            return {
                "s3_key": s3_key,
                "download_url": download_url,
                "s3_bucket": _s3_client.bucket_name
            }
            
        except Exception as e:
            logger.error(f"Error in streaming batch processing: {e}", exc_info=True)
            return None
    
    async def _initialize_excel_structure(self, message: SQSExportMessage) -> Optional[str]:
        # Create initial Excel structure with headers only
        try:
            wb = Workbook()
            sheet_suffix = self._get_sheet_suffix(message)
            
            # Create sheets with headers only (no data yet)
            self._create_sheet_headers(wb, "Top question categories", sheet_suffix, message)
            self._create_sheet_headers(wb, "Detailed category breakdown", sheet_suffix, message) 
            self._create_sheet_headers(wb, "Top asked questions", sheet_suffix, message)
            
            # Remove default sheet
            if len(wb.worksheets) > 1 and wb.worksheets[0].title == "Sheet":
                wb.remove(wb.worksheets[0])
            
            # Upload to S3
            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            s3_key = f"exports/{message.clientId}/{message.eventId}/FAQ_Report_{sheet_suffix}_{timestamp}.xlsx"
            
            if await _s3_client.upload_file(stream, s3_key):
                logger.info(f"Initialized Excel structure on S3: {s3_key}")
                return s3_key
            else:
                logger.error("Failed to upload initial Excel structure")
                return None
                
        except Exception as e:
            logger.error(f"Error initializing Excel structure: {e}", exc_info=True)
            return None
    
    async def _process_all_batches_streaming(self, message: SQSExportMessage, s3_key: str) -> bool:
        # Process all data types in batches: Fetch Batch → Write to Excel → Stream to S3
        try:
            # Process info cards batches
            await self._process_info_cards_batches(message, s3_key)
            
            # Process category distribution batches  
            category_ids = await self._process_category_batches(message, s3_key)
            
            # Process subcategory batches for each category
            await self._process_subcategory_batches(message, s3_key, category_ids)
            
            # Process questions batches
            await self._process_questions_batches(message, s3_key)
            
            logger.info("Successfully processed all batches")
            return True
            
        except Exception as e:
            logger.error(f"Error processing all batches: {e}", exc_info=True)
            return False
    
    async def _process_info_cards_batches(self, message: SQSExportMessage, s3_key: str):
        # Process info cards in batches
        try:
            logger.info("Processing info cards batches")
            
            async for batch in _analytics_client.get_info_cards_batches(
                message.filter, message.pageFilter, BATCH_SIZE
            ):
                # Fetch Batch → Write to Excel → Stream to S3
                await self._add_info_cards_batch_to_excel(s3_key, batch, message)
                logger.info("Processed info cards batch and updated S3")
                
        except Exception as e:
            logger.error(f"Error processing info cards batches: {e}", exc_info=True)
    
    async def _process_category_batches(self, message: SQSExportMessage, s3_key: str) -> List[str]:
        # Process category distribution in batches and return category IDs
        category_ids = []
        
        try:
            logger.info("Processing category distribution batches")
            
            async for batch in _analytics_client.get_category_distribution_batches(
                message.filter, message.pageFilter, BATCH_SIZE
            ):
                # Fetch Batch → Write to Excel → Stream to S3
                await self._add_category_batch_to_excel(s3_key, batch, message)
                
                # Collect category IDs for subcategory processing
                for category in batch.category:
                    category_ids.append(category.id)
                
                logger.info(f"Processed category batch with {len(batch.category)} categories and updated S3")
            
            return category_ids
            
        except Exception as e:
            logger.error(f"Error processing category batches: {e}", exc_info=True)
            return []
    
    async def _process_subcategory_batches(self, message: SQSExportMessage, s3_key: str, category_ids: List[str]):
        # Process subcategory distribution in batches for each category
        try:
            logger.info(f"Processing subcategory batches for {len(category_ids)} categories")
            
            for category_id in category_ids:
                async for batch in _analytics_client.get_subcategory_distribution_batches(
                    category_id, message.filter, message.pageFilter, BATCH_SIZE
                ):
                    # Fetch Batch → Write to Excel → Stream to S3
                    await self._add_subcategory_batch_to_excel(s3_key, batch, message, category_id)
                    logger.info(f"Processed subcategory batch for category {category_id} and updated S3")
                
        except Exception as e:
            logger.error(f"Error processing subcategory batches: {e}", exc_info=True)
    
    async def _process_questions_batches(self, message: SQSExportMessage, s3_key: str):
        # Process questions in batches
        try:
            logger.info("Processing questions batches")
            
            async for batch in _analytics_client.get_top_questions_batches(
                message.filter, message.pageFilter, BATCH_SIZE
            ):
                # Fetch Batch → Write to Excel → Stream to S3
                await self._add_questions_batch_to_excel(s3_key, batch, message)
                logger.info(f"Processed questions batch with {len(batch.topQuestions)} questions and updated S3")
                
        except Exception as e:
            logger.error(f"Error processing questions batches: {e}", exc_info=True)
    
    async def _add_info_cards_batch_to_excel(self, s3_key: str, batch, message: SQSExportMessage):
        # Add info cards batch to Excel and upload to S3
        try:
            # Download → Modify → Upload cycle
            excel_stream = await _s3_client.download_file(s3_key)
            wb = load_workbook(excel_stream)
            
            # Process info cards data (store for frequency calculations)
            # Info cards are typically used for calculations, not direct Excel rows
            
            # Save back to S3
            updated_stream = BytesIO()
            wb.save(updated_stream)
            updated_stream.seek(0)
            await _s3_client.upload_file(updated_stream, s3_key)
            
        except Exception as e:
            logger.error(f"Error adding info cards batch to Excel: {e}", exc_info=True)
    
    async def _add_category_batch_to_excel(self, s3_key: str, batch, message: SQSExportMessage):
        # Add category batch to Excel and upload to S3
        try:
            # Download → Modify → Upload cycle
            excel_stream = await _s3_client.download_file(s3_key)
            wb = load_workbook(excel_stream)
            
            sheet_suffix = self._get_sheet_suffix(message)
            ws = wb[f"Top question categories - {sheet_suffix}"]
            
            # Find next available row
            next_row = self._find_next_available_row(ws, start_row=9)
            
            # Add category data
            for category in batch.category:
                # For now, using placeholder frequency calculation
                frequency = int(1000 * (category.distribution / 100))  # Will be improved with info cards
                
                ws.cell(row=next_row, column=1, value=category.category)
                ws.cell(row=next_row, column=2, value=frequency)
                ws.cell(row=next_row, column=3, value=f"{category.distribution:.2f}%")
                
                trend_symbol = "▲" if category.direction == "INCREASING" else "▼"
                ws.cell(row=next_row, column=4, value=f"{trend_symbol} {abs(category.trend):.0f}%")
                ws.cell(row=next_row, column=5, value="View details ↗")
                next_row += 1
            
            # Save back to S3
            updated_stream = BytesIO()
            wb.save(updated_stream)
            updated_stream.seek(0)
            await _s3_client.upload_file(updated_stream, s3_key)
            
        except Exception as e:
            logger.error(f"Error adding category batch to Excel: {e}", exc_info=True)
    
    async def _add_subcategory_batch_to_excel(self, s3_key: str, batch, message: SQSExportMessage, category_id: str):
        # Add subcategory batch to Excel and upload to S3
        try:
            # Download → Modify → Upload cycle
            excel_stream = await _s3_client.download_file(s3_key)
            wb = load_workbook(excel_stream)
            
            sheet_suffix = self._get_sheet_suffix(message)
            ws = wb[f"Detailed category breakdown - {sheet_suffix}"]
            
            # Find next available row
            next_row = self._find_next_available_row(ws, start_row=9)
            
            # Add subcategory data
            for subcategory in batch.subCategory:
                frequency = int(1000 * (subcategory.distribution / 100))  # Placeholder calculation
                
                ws.cell(row=next_row, column=1, value=f"→ {subcategory.subCategory}")
                ws.cell(row=next_row, column=2, value=frequency)
                ws.cell(row=next_row, column=3, value=f"{subcategory.distribution:.2f}%")
                
                trend_symbol = "▲" if subcategory.direction == "INCREASING" else "▼"
                ws.cell(row=next_row, column=4, value=f"{trend_symbol} {abs(subcategory.trend):.0f}%")
                ws.cell(row=next_row, column=5, value="View details ↗")
                next_row += 1
            
            # Save back to S3
            updated_stream = BytesIO()
            wb.save(updated_stream)
            updated_stream.seek(0)
            await _s3_client.upload_file(updated_stream, s3_key)
            
        except Exception as e:
            logger.error(f"Error adding subcategory batch to Excel: {e}", exc_info=True)
    
    async def _add_questions_batch_to_excel(self, s3_key: str, batch, message: SQSExportMessage):
        # Add questions batch to Excel and upload to S3
        try:
            # Download → Modify → Upload cycle
            excel_stream = await _s3_client.download_file(s3_key)
            wb = load_workbook(excel_stream)
            
            sheet_suffix = self._get_sheet_suffix(message)
            ws = wb[f"Top asked questions - {sheet_suffix}"]
            
            # Find next available row
            next_row = self._find_next_available_row(ws, start_row=10)
            
            # Add questions data
            for question in batch.topQuestions:
                ws.cell(row=next_row, column=1, value=question.question)
                ws.cell(row=next_row, column=2, value=question.frequency)
                ws.cell(row=next_row, column=3, value="View details ↗")
                next_row += 1
            
            # Save back to S3
            updated_stream = BytesIO()
            wb.save(updated_stream)
            updated_stream.seek(0)
            await _s3_client.upload_file(updated_stream, s3_key)
            
        except Exception as e:
            logger.error(f"Error adding questions batch to Excel: {e}", exc_info=True)
    
    def _create_sheet_headers(self, wb: Workbook, sheet_type: str, sheet_suffix: str, message: SQSExportMessage):
        # Create sheet with headers only
        ws = wb.create_sheet(f"{sheet_type} - {sheet_suffix}")
        
        if sheet_type == "Top question categories":
            ws['A1'] = "Top question categories"
            ws['A2'] = f"Date range - {self._get_date_range_string(message.pageFilter)}"
            ws['A4'] = "Filters applied -"
            ws['A5'] = "Users : (All, single user, or comma separated)"
            ws['A6'] = "Initiated from : (All, single source, or comma separated)"
            
            headers = ["Category", "Frequency (Questions asked)", "Distribution", "Trend", "Link"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=8, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        elif sheet_type == "Detailed category breakdown":
            ws['A1'] = "Detailed category breakdown"
            ws['A2'] = f"Date range - {self._get_date_range_string(message.pageFilter)}"
            ws['A4'] = "Filters applied -"
            ws['A5'] = "Users : (All, single user, or comma separated)"
            ws['A6'] = "Initiated from : (All, single source, or comma separated)"
            
            headers = ["Subcategory", "Frequency (Questions asked)", "Distribution", "Trend", "Link"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=8, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        elif sheet_type == "Top asked questions":
            ws['A1'] = "Top asked questions"
            ws['A2'] = f"Date range - {self._get_date_range_string(message.pageFilter)}"
            ws['A3'] = "💡 Questions that are similar to each other have been grouped under a single FAQ"
            ws['A5'] = "Filters applied -"
            ws['A6'] = "Users : (All, single user, or comma separated)"
            ws['A7'] = "Initiated from : (All, single source, or comma separated)"
            
            headers = ["Question", "Frequency (Questions asked)", "Link"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=9, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    def _find_next_available_row(self, ws, start_row: int) -> int:
        # Find the next available row for data insertion
        row = start_row
        while ws.cell(row=row, column=1).value is not None:
            row += 1
        return row
    
    def _get_sheet_suffix(self, message: SQSExportMessage) -> str:
        # Get sheet suffix based on filter data
        try:
            if message.filter and message.filter.conditions:
                status_condition = message.filter.conditions.get("status")
                if status_condition and status_condition.data:
                    data = status_condition.data
                    if "ANSWERED#@#NO_INFORMATION#@#PARTIAL" in data:
                        return "All"
                    elif "ANSWERED#@#PARTIAL" in data:
                        return "Answered"
                    elif "NO_INFORMATION" in data:
                        return "Unanswered"
            return "All"
        except Exception:
            return "All"
    
    def _get_date_range_string(self, page_filter) -> str:
        # Extract date range from page filter
        try:
            if page_filter and page_filter.conditions:
                meta_created = page_filter.conditions.get("meta.created")
                if meta_created and meta_created.data:
                    timestamps = meta_created.data.split("#@#")
                    if len(timestamps) == 2:
                        start_ts = int(timestamps[0]) / 1000
                        end_ts = int(timestamps[1]) / 1000
                        
                        start_date = datetime.fromtimestamp(start_ts).strftime("%b %d, %Y")
                        end_date = datetime.fromtimestamp(end_ts).strftime("%b %d, %Y")
                        
                        return f"{start_date} to {end_date}"
            
            return "Date range not specified"
            
        except Exception as e:
            logger.error(f"Error parsing date range: {e}")
            return "Date range not specified" 