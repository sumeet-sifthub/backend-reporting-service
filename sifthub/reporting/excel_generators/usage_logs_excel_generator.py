import time
from io import BytesIO
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

from sifthub.reporting.models.export_models import SQSExportRequest
from sifthub.reporting.services.usage_logs_analytics_client import UsageLogsAnalyticsClient
from sifthub.configs.constants import BATCH_SIZE
from sifthub.datastores.document.s3.s3_client import S3Client
from sifthub.utils.logger import setup_logger

logger = setup_logger(__name__)

# Module-level clients
_analytics_client = UsageLogsAnalyticsClient()
_s3_client = S3Client()


class UsageLogsExcelGenerator:
    # Excel generator implementing true streaming batch processing for UsageLogs
    
    async def generate_excel_streaming(self, message: SQSExportRequest) -> Optional[Dict[str, str]]:
        # Implement the exact flow: Fetch Batch → Write to Excel → Stream to S3 → Repeat
        try:
            logger.info(f"Starting UsageLogs streaming batch processing for event: {message.eventId}, type: {message.type}")
            
            # Step 1: Initialize Excel structure and upload to S3
            s3_key = await self._initialize_excel_structure(message)
            if not s3_key:
                logger.error("Failed to initialize Excel structure")
                return None
            
            # Step 2: Process all batches sequentially
            success = await self._process_all_batches_streaming(message, s3_key)
            if not success:
                logger.error("Failed to process all batches")
                return None
            
            # Step 3: Generate presigned URL (after all batching is complete)
            download_url = await _s3_client.generate_presigned_url(s3_key)
            if not download_url:
                logger.error("Failed to generate presigned URL")
                return None
            
            logger.info(f"Successfully completed UsageLogs streaming batch processing for event: {message.eventId}")
            
            # Step 4: Return result for Firebase notification
            return {
                "s3_key": s3_key,
                "download_url": download_url,
                "s3_bucket": _s3_client.bucket_name
            }
            
        except Exception as e:
            logger.error(f"Error in UsageLogs streaming batch processing: {e}", exc_info=True)
            return None
    
    async def _initialize_excel_structure(self, message: SQSExportRequest) -> Optional[str]:
        # Create initial Excel structure with headers only
        try:
            wb = Workbook()
            
            # Create sheets with headers based on type
            if message.type == "answer":
                self._create_answer_sheet_headers(wb, message)
            elif message.type == "autofill":
                self._create_autofill_sheet_headers(wb, message)
            elif message.type == "AITeammate":
                self._create_aiteammate_sheet_headers(wb, message)
            else:
                logger.error(f"Unsupported UsageLogs type: {message.type}")
                return None
            
            # Remove default sheet
            if len(wb.worksheets) > 1 and wb.worksheets[0].title == "Sheet":
                wb.remove(wb.worksheets[0])
            
            # Upload to S3
            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            date_range = self._get_date_range_string(message.pageFilter)
            s3_key = f"exports/{message.clientId}/{message.eventId}/{message.type.title()}_Usage_logs_{date_range}_{timestamp}.xlsx"
            
            if await _s3_client.upload_file(stream, s3_key):
                logger.info(f"Initialized UsageLogs Excel structure on S3: {s3_key}")
                return s3_key
            else:
                logger.error("Failed to upload initial Excel structure")
                return None
                
        except Exception as e:
            logger.error(f"Error initializing Excel structure: {e}", exc_info=True)
            return None
    
    async def _process_all_batches_streaming(self, message: SQSExportRequest, s3_key: str) -> bool:
        # Process all data types in batches: Fetch Batch → Write to Excel → Stream to S3
        try:
            if message.type == "answer":
                await self._process_answer_batches(message, s3_key)
            elif message.type == "autofill":
                await self._process_autofill_batches(message, s3_key)
            elif message.type == "AITeammate":
                await self._process_aiteammate_batches(message, s3_key)
            
            logger.info("Successfully processed all UsageLogs batches")
            return True
            
        except Exception as e:
            logger.error(f"Error processing all UsageLogs batches: {e}", exc_info=True)
            return False
    
    # Answer processing methods
    async def _process_answer_batches(self, message: SQSExportRequest, s3_key: str):
        # Process answer logs and stats
        try:
            logger.info("Processing answer batches")
            
            # Process logs in batches
            async for batch in _analytics_client.get_answer_logs_batches(
                message.filter, message.pageFilter, BATCH_SIZE
            ):
                await self._add_answer_logs_batch_to_excel(s3_key, batch, message)
                logger.info(f"Processed answer logs batch with {len(batch.data)} records and updated S3")
            
            # Process stats (single call)
            stats = await _analytics_client.get_answer_stats(message.filter, message.pageFilter)
            if stats:
                await self._add_answer_stats_to_excel(s3_key, stats, message)
                logger.info("Processed answer stats and updated S3")
                
        except Exception as e:
            logger.error(f"Error processing answer batches: {e}", exc_info=True)
    
    async def _add_answer_logs_batch_to_excel(self, s3_key: str, batch, message: SQSExportRequest):
        # Add answer logs batch to Excel and upload to S3
        try:
            # Download → Modify → Upload cycle
            excel_stream = await _s3_client.download_file(s3_key)
            wb = load_workbook(excel_stream)
            
            ws = wb["Answer Usage logs - Logs"]
            
            # Find next available row
            next_row = self._find_next_available_row(ws, start_row=9)
            
            # Add logs data
            for log in batch.data:
                ws.cell(row=next_row, column=1, value=log.question)
                ws.cell(row=next_row, column=2, value=log.userInstruction)
                ws.cell(row=next_row, column=3, value=log.answer or "")
                
                # Sources - join URLs
                sources_urls = ", ".join([src.url for src in log.sources]) if log.sources else ""
                ws.cell(row=next_row, column=4, value=sources_urls)
                
                ws.cell(row=next_row, column=5, value=log.status)
                
                # Date from meta.created
                created_date = datetime.fromtimestamp(log.meta.created / 1000).strftime("%b %d, %Y")
                ws.cell(row=next_row, column=6, value=created_date)
                
                ws.cell(row=next_row, column=7, value=log.meta.createdBy.fullName)
                ws.cell(row=next_row, column=8, value=log.initiatedFrom)
                ws.cell(row=next_row, column=9, value=log.txConsumed)
                
                next_row += 1
            
            # Save back to S3
            updated_stream = BytesIO()
            wb.save(updated_stream)
            updated_stream.seek(0)
            await _s3_client.upload_file(updated_stream, s3_key)
            
        except Exception as e:
            logger.error(f"Error adding answer logs batch to Excel: {e}", exc_info=True)
    
    async def _add_answer_stats_to_excel(self, s3_key: str, stats, message: SQSExportRequest):
        # Add answer stats to Excel and upload to S3
        try:
            # Download → Modify → Upload cycle
            excel_stream = await _s3_client.download_file(s3_key)
            wb = load_workbook(excel_stream)
            
            ws = wb["Answer Usage logs - Summary"]
            
            # Add stats data
            ws.cell(row=7, column=2, value=stats.data.total)  # Total questions asked
            ws.cell(row=8, column=2, value=stats.data.answered)  # Total questions answered
            ws.cell(row=9, column=2, value=stats.data.noInformation)  # No information found
            ws.cell(row=10, column=2, value=stats.data.txConsumed)  # Transactions consumed
            
            # Save back to S3
            updated_stream = BytesIO()
            wb.save(updated_stream)
            updated_stream.seek(0)
            await _s3_client.upload_file(updated_stream, s3_key)
            
        except Exception as e:
            logger.error(f"Error adding answer stats to Excel: {e}", exc_info=True)
    
    # Autofill processing methods
    async def _process_autofill_batches(self, message: SQSExportRequest, s3_key: str):
        # Process autofill logs and stats
        try:
            logger.info("Processing autofill batches")
            
            # Process logs in batches
            async for batch in _analytics_client.get_autofill_logs_batches(
                message.filter, message.pageFilter, BATCH_SIZE
            ):
                await self._add_autofill_logs_batch_to_excel(s3_key, batch, message)
                logger.info(f"Processed autofill logs batch with {len(batch.data)} records and updated S3")
            
            # Process stats (single call)
            stats = await _analytics_client.get_autofill_stats(message.filter, message.pageFilter)
            if stats:
                await self._add_autofill_stats_to_excel(s3_key, stats, message)
                logger.info("Processed autofill stats and updated S3")
                
        except Exception as e:
            logger.error(f"Error processing autofill batches: {e}", exc_info=True)
    
    async def _add_autofill_logs_batch_to_excel(self, s3_key: str, batch, message: SQSExportRequest):
        # Add autofill logs batch to Excel and upload to S3 (same structure as answer)
        try:
            # Download → Modify → Upload cycle
            excel_stream = await _s3_client.download_file(s3_key)
            wb = load_workbook(excel_stream)
            
            ws = wb["Autofill Usage logs - Logs"]
            
            # Find next available row
            next_row = self._find_next_available_row(ws, start_row=9)
            
            # Add logs data (same structure as answer)
            for log in batch.data:
                ws.cell(row=next_row, column=1, value=log.question)
                ws.cell(row=next_row, column=2, value=log.userInstruction)
                ws.cell(row=next_row, column=3, value=log.answer or "")
                
                # Sources - join URLs
                sources_urls = ", ".join([src.url for src in log.sources]) if log.sources else ""
                ws.cell(row=next_row, column=4, value=sources_urls)
                
                ws.cell(row=next_row, column=5, value=log.status)
                
                # Date from meta.created
                created_date = datetime.fromtimestamp(log.meta.created / 1000).strftime("%b %d, %Y")
                ws.cell(row=next_row, column=6, value=created_date)
                
                ws.cell(row=next_row, column=7, value=log.meta.createdBy.fullName)
                ws.cell(row=next_row, column=8, value=log.initiatedFrom)
                ws.cell(row=next_row, column=9, value=log.txConsumed)
                
                next_row += 1
            
            # Save back to S3
            updated_stream = BytesIO()
            wb.save(updated_stream)
            updated_stream.seek(0)
            await _s3_client.upload_file(updated_stream, s3_key)
            
        except Exception as e:
            logger.error(f"Error adding autofill logs batch to Excel: {e}", exc_info=True)
    
    async def _add_autofill_stats_to_excel(self, s3_key: str, stats, message: SQSExportRequest):
        # Add autofill stats to Excel and upload to S3
        try:
            # Download → Modify → Upload cycle
            excel_stream = await _s3_client.download_file(s3_key)
            wb = load_workbook(excel_stream)
            
            ws = wb["Autofill Usage logs - Summary"]
            
            # Add stats data
            ws.cell(row=7, column=2, value=stats.data.totalRuns)  # Autofill runs
            ws.cell(row=8, column=2, value=stats.data.totalDocuments)  # Documents autofilled
            ws.cell(row=9, column=2, value=stats.data.totalQuestions)  # Total questions
            ws.cell(row=10, column=2, value=stats.data.totalQuestionsAnswered)  # Questions answered
            ws.cell(row=11, column=2, value=stats.data.averageResponseTime)  # Average response time
            
            # Save back to S3
            updated_stream = BytesIO()
            wb.save(updated_stream)
            updated_stream.seek(0)
            await _s3_client.upload_file(updated_stream, s3_key)
            
        except Exception as e:
            logger.error(f"Error adding autofill stats to Excel: {e}", exc_info=True)
    
    # AITeammate processing methods
    async def _process_aiteammate_batches(self, message: SQSExportRequest, s3_key: str):
        # Process AITeammate logs and stats
        try:
            logger.info("Processing AITeammate batches")
            
            # Process logs in batches
            async for batch in _analytics_client.get_teammate_logs_batches(
                message.filter, message.pageFilter, BATCH_SIZE
            ):
                await self._add_aiteammate_logs_batch_to_excel(s3_key, batch, message)
                logger.info(f"Processed AITeammate logs batch with {len(batch.data)} records and updated S3")
            
            # Process stats (single call)
            stats = await _analytics_client.get_teammate_stats(message.filter, message.pageFilter)
            if stats:
                await self._add_aiteammate_stats_to_excel(s3_key, stats, message)
                logger.info("Processed AITeammate stats and updated S3")
                
        except Exception as e:
            logger.error(f"Error processing AITeammate batches: {e}", exc_info=True)
    
    async def _add_aiteammate_logs_batch_to_excel(self, s3_key: str, batch, message: SQSExportRequest):
        # Add AITeammate logs batch to Excel and upload to S3
        try:
            # Download → Modify → Upload cycle
            excel_stream = await _s3_client.download_file(s3_key)
            wb = load_workbook(excel_stream)
            
            ws = wb["AITeammate Usage logs - Logs"]
            
            # Find next available row
            next_row = self._find_next_available_row(ws, start_row=9)
            
            # Add logs data (different structure for AITeammate)
            for log in batch.data:
                ws.cell(row=next_row, column=1, value=log.title)  # Conversations
                
                # Date from meta.created
                created_date = datetime.fromtimestamp(log.meta.created / 1000).strftime("%b %d, %Y")
                ws.cell(row=next_row, column=2, value=created_date)
                
                ws.cell(row=next_row, column=3, value=log.meta.createdBy.fullName)  # Owner
                ws.cell(row=next_row, column=4, value=log.threadCount)  # No. of Turns
                ws.cell(row=next_row, column=5, value=log.averageTime)  # Response time per response
                ws.cell(row=next_row, column=6, value=log.txConsumed)  # Transactions consumed
                
                next_row += 1
            
            # Save back to S3
            updated_stream = BytesIO()
            wb.save(updated_stream)
            updated_stream.seek(0)
            await _s3_client.upload_file(updated_stream, s3_key)
            
        except Exception as e:
            logger.error(f"Error adding AITeammate logs batch to Excel: {e}", exc_info=True)
    
    async def _add_aiteammate_stats_to_excel(self, s3_key: str, stats, message: SQSExportRequest):
        # Add AITeammate stats to Excel and upload to S3
        try:
            # Download → Modify → Upload cycle
            excel_stream = await _s3_client.download_file(s3_key)
            wb = load_workbook(excel_stream)
            
            ws = wb["AITeammate Usage logs - Summary"]
            
            # Add stats data
            ws.cell(row=7, column=2, value=stats.data.threadCount)  # Total Conversations
            ws.cell(row=8, column=2, value=stats.data.averageTime)  # Average response time
            ws.cell(row=9, column=2, value=stats.data.txConsumed)  # Transactions consumed
            
            # Save back to S3
            updated_stream = BytesIO()
            wb.save(updated_stream)
            updated_stream.seek(0)
            await _s3_client.upload_file(updated_stream, s3_key)
            
        except Exception as e:
            logger.error(f"Error adding AITeammate stats to Excel: {e}", exc_info=True)
    
    # Sheet creation methods
    def _create_answer_sheet_headers(self, wb: Workbook, message: SQSExportRequest):
        # Create Answer sheets with headers
        date_range = self._get_date_range_string(message.pageFilter)
        
        # Logs sheet
        ws_logs = wb.create_sheet("Answer Usage logs - Logs")
        ws_logs['A1'] = "Answer Usage logs"
        ws_logs['A2'] = f"Date range - {date_range}"
        ws_logs['A4'] = "Filters applied -"
        ws_logs['A5'] = "Users : (All, single user, or comma separated)"
        ws_logs['A6'] = "Status: (All, single or comma separated)"
        ws_logs['A7'] = "Initiated from : (All, single source, or comma separated)"
        
        # Column headers for logs
        headers = ["Question", "Instruction", "Answer", "Sources", "Status", "Date", "User", "Initiated from", "Transactions consumed"]
        for col, header in enumerate(headers, 1):
            cell = ws_logs.cell(row=8, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Summary sheet
        ws_summary = wb.create_sheet("Answer Usage logs - Summary")
        ws_summary['A1'] = "Answer Usage logs - Summary"
        ws_summary['A2'] = f"Date range - {date_range}"
        ws_summary['A4'] = "Filters applied -"
        ws_summary['A5'] = "Users : (All, single user, or comma separated)"
        ws_summary['A6'] = "Status: (All, single or comma separated)"
        ws_summary['A7'] = "Initiated from : (All, single source, or comma separated)"
        
        # Metric headers
        ws_summary.cell(row=6, column=1, value="Metric").font = Font(bold=True)
        ws_summary.cell(row=6, column=2, value="Value").font = Font(bold=True)
        ws_summary.cell(row=6, column=1).fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        ws_summary.cell(row=6, column=2).fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        ws_summary.cell(row=7, column=1, value="Total questions asked")
        ws_summary.cell(row=8, column=1, value="Total questions answered")
        ws_summary.cell(row=9, column=1, value="No information found")
        ws_summary.cell(row=10, column=1, value="Transactions consumed")
    
    def _create_autofill_sheet_headers(self, wb: Workbook, message: SQSExportRequest):
        # Create Autofill sheets with headers (same logs structure as Answer)
        date_range = self._get_date_range_string(message.pageFilter)
        
        # Logs sheet
        ws_logs = wb.create_sheet("Autofill Usage logs - Logs")
        ws_logs['A1'] = "Autofill Usage logs"
        ws_logs['A2'] = f"Date range - {date_range}"
        ws_logs['A4'] = "Filters applied -"
        ws_logs['A5'] = "Users : (All, single user, or comma separated)"
        ws_logs['A6'] = "Status: (All, single or comma separated)"
        ws_logs['A7'] = "Initiated from : (All, single source, or comma separated)"
        
        # Column headers for logs (same as Answer)
        headers = ["Question", "Instruction", "Answer", "Sources", "Status", "Date", "User", "Initiated from", "Transactions consumed"]
        for col, header in enumerate(headers, 1):
            cell = ws_logs.cell(row=8, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Summary sheet
        ws_summary = wb.create_sheet("Autofill Usage logs - Summary")
        ws_summary['A1'] = "Autofill Usage logs - Summary"
        ws_summary['A2'] = f"Date range - {date_range}"
        ws_summary['A4'] = "Filters applied -"
        ws_summary['A5'] = "Users : (All, single user, or comma separated)"
        ws_summary['A6'] = "Status: (All, single or comma separated)"
        ws_summary['A7'] = "Initiated from : (All, single source, or comma separated)"
        
        # Metric headers
        ws_summary.cell(row=6, column=1, value="Metric").font = Font(bold=True)
        ws_summary.cell(row=6, column=2, value="Value").font = Font(bold=True)
        ws_summary.cell(row=6, column=1).fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        ws_summary.cell(row=6, column=2).fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        ws_summary.cell(row=7, column=1, value="Autofill runs")
        ws_summary.cell(row=8, column=1, value="Documents autofilled")
        ws_summary.cell(row=9, column=1, value="Total questions")
        ws_summary.cell(row=10, column=1, value="Questions answered")
        ws_summary.cell(row=11, column=1, value="Average response time")
    
    def _create_aiteammate_sheet_headers(self, wb: Workbook, message: SQSExportRequest):
        # Create AITeammate sheets with headers (different structure)
        date_range = self._get_date_range_string(message.pageFilter)
        
        # Logs sheet
        ws_logs = wb.create_sheet("AITeammate Usage logs - Logs")
        ws_logs['A1'] = "AITeammate Usage logs"
        ws_logs['A2'] = f"Date range - {date_range}"
        ws_logs['A4'] = "Filters applied -"
        ws_logs['A5'] = "Users : (All, single user, or comma separated)"
        ws_logs['A6'] = "Status: (All, single or comma separated)"
        ws_logs['A7'] = "Initiated from : (All, single source, or comma separated)"
        
        # Column headers for logs (different for AITeammate)
        headers = ["Conversations", "Date", "Owner", "No. of Turns", "Response time per response", "Transactions consumed"]
        for col, header in enumerate(headers, 1):
            cell = ws_logs.cell(row=8, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Summary sheet
        ws_summary = wb.create_sheet("AITeammate Usage logs - Summary")
        ws_summary['A1'] = "AITeammate Usage logs - Summary"
        ws_summary['A2'] = f"Date range - {date_range}"
        ws_summary['A4'] = "Filters applied -"
        ws_summary['A5'] = "Users : (All, single user, or comma separated)"
        ws_summary['A6'] = "Status: (All, single or comma separated)"
        ws_summary['A7'] = "Initiated from : (All, single source, or comma separated)"
        
        # Metric headers
        ws_summary.cell(row=6, column=1, value="Metric").font = Font(bold=True)
        ws_summary.cell(row=6, column=2, value="Value").font = Font(bold=True)
        ws_summary.cell(row=6, column=1).fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        ws_summary.cell(row=6, column=2).fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        ws_summary.cell(row=7, column=1, value="Total Conversations")
        ws_summary.cell(row=8, column=1, value="Average response time")
        ws_summary.cell(row=9, column=1, value="Transactions consumed")
    
    def _find_next_available_row(self, ws, start_row: int) -> int:
        # Find the next available row for data insertion
        row = start_row
        while ws.cell(row=row, column=1).value is not None:
            row += 1
        return row
    
    def _get_date_range_string(self, page_filter) -> str:
        # Extract date range from page filter
        try:
            if page_filter and page_filter.conditions:
                meta_created = page_filter.conditions.get("meta.created")
                if meta_created and meta_created.data:
                    timestamps = meta_created.data.split("#@#")
                    if len(timestamps) == 2:
                        start_ts = int(timestamps[0]) / 1000
                        end_ts = int(timestamps[1]) / 1000
                        
                        start_date = datetime.fromtimestamp(start_ts).strftime("%b %d, %Y")
                        end_date = datetime.fromtimestamp(end_ts).strftime("%b %d, %Y")
                        
                        return f"{start_date} to {end_date}"
            
            return "Jan 03, 2025 to Feb 03, 2024"  # Default fallback
            
        except Exception as e:
            logger.error(f"Error parsing date range: {e}")
            return "Jan 03, 2025 to Feb 03, 2024" 